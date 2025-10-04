from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_from_directory, send_file
from flask_socketio import SocketIO, emit, join_room, leave_room
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
import sqlite3
import os
import string
import random
import secrets
from datetime import datetime
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

APP_DIR = os.path.dirname(__file__)
DB_PATH = os.path.join(APP_DIR, 'db.sqlite3')
UPLOAD_FOLDER = os.path.join(APP_DIR, 'static', 'uploads', 'profile_photos')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
BRASILIA_TZ = pytz.timezone('America/Sao_Paulo')

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SESSION_SECRET', 'troque-esta-chave-por-uma-segura')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024
socketio = SocketIO(app, 
                    cors_allowed_origins="*",
                    async_mode='threading',
                    ping_timeout=60,
                    ping_interval=25,
                    engineio_logger=False,
                    logger=False,
                    transports=['websocket', 'polling'],
                    always_connect=True)

_system_config_cache = None
_config_cache_time = None
CONFIG_CACHE_DURATION = 300

room_users = {}

def get_cached_system_config():
    """Retorna configurações do sistema com cache de 5 minutos"""
    global _system_config_cache, _config_cache_time
    
    now = datetime.now()
    if _system_config_cache and _config_cache_time and (now - _config_cache_time).total_seconds() < CONFIG_CACHE_DURATION:
        return _system_config_cache
    
    conn = get_db()
    try:
        config = conn.execute('SELECT platform_name, support_email, support_phone FROM sistema_config WHERE id = 1').fetchone()
        if config:
            _system_config_cache = {
                'platform_name': config['platform_name'],
                'support_email': config['support_email'],
                'support_phone': config['support_phone']
            }
        else:
            _system_config_cache = {
                'platform_name': 'MedConnect',
                'support_email': 'suporte@medconnect.com',
                'support_phone': '(11) 3000-0000'
            }
        _config_cache_time = now
        return _system_config_cache
    except:
        return {
            'platform_name': 'MedConnect',
            'support_email': 'suporte@medconnect.com',
            'support_phone': '(11) 3000-0000'
        }
    finally:
        conn.close()

@app.context_processor
def inject_config():
    """Injeta configurações do sistema em todos os templates"""
    result = {
        'system_config': get_cached_system_config(),
        'user_profile_photo': None
    }
    
    if 'user_id' in session:
        conn = get_db()
        try:
            user = conn.execute('SELECT profile_photo FROM users WHERE id = ?', (session['user_id'],)).fetchone()
            if user and user['profile_photo']:
                result['user_profile_photo'] = user['profile_photo']
        except:
            pass
        finally:
            conn.close()
    
    return result

def get_brasilia_time():
    """Retorna datetime atual no timezone de Brasília"""
    return datetime.now(BRASILIA_TZ)

def format_date_br(date_string):
    """Formata data para dd/mm/aa"""
    if not date_string:
        return ''
    try:
        if isinstance(date_string, str):
            date_string = date_string.split('.')[0]
            if 'T' in date_string or ' ' in date_string:
                if 'T' in date_string:
                    dt = datetime.fromisoformat(date_string.replace('Z', '+00:00'))
                else:
                    dt = datetime.strptime(date_string, '%Y-%m-%d %H:%M:%S')
                if dt.tzinfo is None:
                    utc_dt = pytz.UTC.localize(dt)
                else:
                    utc_dt = dt
                br_dt = utc_dt.astimezone(BRASILIA_TZ)
                return br_dt.strftime('%d/%m/%y')
            else:
                dt = datetime.strptime(date_string, '%Y-%m-%d')
                return dt.strftime('%d/%m/%y')
        else:
            dt = date_string
            if dt.tzinfo is None:
                return dt.strftime('%d/%m/%y')
            else:
                br_dt = dt.astimezone(BRASILIA_TZ)
                return br_dt.strftime('%d/%m/%y')
    except Exception as e:
        return date_string

def format_datetime_br(date_string):
    """Formata data e hora para dd/mm/aa HH:MM"""
    if not date_string:
        return ''
    try:
        if isinstance(date_string, str):
            date_string = date_string.split('.')[0]
            if 'T' in date_string:
                dt = datetime.fromisoformat(date_string.replace('Z', '+00:00'))
                if dt.tzinfo is None:
                    utc_dt = pytz.UTC.localize(dt)
                else:
                    utc_dt = dt
                br_dt = utc_dt.astimezone(BRASILIA_TZ)
                return br_dt.strftime('%d/%m/%y %H:%M')
            elif ' ' in date_string:
                dt = datetime.strptime(date_string, '%Y-%m-%d %H:%M:%S')
                if dt.tzinfo is None:
                    utc_dt = pytz.UTC.localize(dt)
                else:
                    utc_dt = dt
                br_dt = utc_dt.astimezone(BRASILIA_TZ)
                return br_dt.strftime('%d/%m/%y %H:%M')
            else:
                dt = datetime.strptime(date_string, '%Y-%m-%d')
                return dt.strftime('%d/%m/%y 00:00')
        else:
            dt = date_string
            if dt.tzinfo is None:
                return dt.strftime('%d/%m/%y %H:%M')
            else:
                br_dt = dt.astimezone(BRASILIA_TZ)
                return br_dt.strftime('%d/%m/%y %H:%M')
    except Exception as e:
        return date_string

app.jinja_env.filters['date_br'] = format_date_br
app.jinja_env.filters['datetime_br'] = format_datetime_br

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=20.0, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA synchronous=NORMAL')
    conn.execute('PRAGMA cache_size=-64000')
    conn.execute('PRAGMA temp_store=MEMORY')
    conn.execute('PRAGMA mmap_size=134217728')
    return conn

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    cur.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE NOT NULL,
        senha_hash TEXT NOT NULL,
        nome TEXT NOT NULL,
        tipo TEXT NOT NULL CHECK(tipo IN ('admin', 'medico', 'paciente')),
        criado_em DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS medicos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL UNIQUE,
        crm TEXT,
        especialidade TEXT,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )''')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS pacientes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL UNIQUE,
        medico_id INTEGER,
        data_nascimento DATE,
        telefone TEXT,
        FOREIGN KEY (user_id) REFERENCES users(id),
        FOREIGN KEY (medico_id) REFERENCES medicos(id)
    )''')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS salas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT UNIQUE NOT NULL,
        medico_id INTEGER NOT NULL,
        paciente_id INTEGER NOT NULL,
        titulo TEXT,
        ativa BOOLEAN DEFAULT 1,
        criado_em DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (medico_id) REFERENCES medicos(id),
        FOREIGN KEY (paciente_id) REFERENCES pacientes(id),
        UNIQUE(medico_id, paciente_id)
    )''')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS evolucoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sala_id INTEGER,
        medico_id INTEGER NOT NULL,
        paciente_id INTEGER NOT NULL,
        data DATETIME DEFAULT CURRENT_TIMESTAMP,
        anotacoes TEXT,
        diagnostico TEXT,
        prescricao TEXT,
        FOREIGN KEY (sala_id) REFERENCES salas(id),
        FOREIGN KEY (medico_id) REFERENCES medicos(id),
        FOREIGN KEY (paciente_id) REFERENCES pacientes(id)
    )''')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS faturamento_config (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        medico_id INTEGER NOT NULL UNIQUE,
        valor_consulta REAL DEFAULT 150.00,
        porcentagem_medico REAL DEFAULT 70.00,
        porcentagem_admin REAL DEFAULT 30.00,
        FOREIGN KEY (medico_id) REFERENCES medicos(id)
    )''')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS consultas_realizadas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        evolucao_id INTEGER,
        sala_id INTEGER NOT NULL,
        medico_id INTEGER NOT NULL,
        paciente_id INTEGER NOT NULL,
        data DATETIME DEFAULT CURRENT_TIMESTAMP,
        data_dia DATE DEFAULT (DATE(CURRENT_TIMESTAMP)),
        valor_total REAL NOT NULL,
        valor_medico REAL NOT NULL,
        valor_admin REAL NOT NULL,
        status TEXT DEFAULT 'pendente' CHECK(status IN ('pendente', 'pago')),
        FOREIGN KEY (evolucao_id) REFERENCES evolucoes(id),
        FOREIGN KEY (sala_id) REFERENCES salas(id),
        FOREIGN KEY (medico_id) REFERENCES medicos(id),
        FOREIGN KEY (paciente_id) REFERENCES pacientes(id),
        UNIQUE(sala_id, data_dia)
    )''')
    
    # Migração: verificar se data_dia existe e adicionar se necessário
    cur.execute("PRAGMA table_info(consultas_realizadas)")
    columns = [col[1] for col in cur.fetchall()]
    
    if 'data_dia' not in columns:
        # Precisamos migrar o banco existente
        # Verificar se há dados na tabela antiga
        cur.execute("SELECT COUNT(*) FROM consultas_realizadas")
        tem_dados = cur.fetchone()[0] > 0
        
        if tem_dados:
            # Criar tabela temporária com nova estrutura
            cur.execute('''CREATE TABLE consultas_realizadas_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                evolucao_id INTEGER,
                sala_id INTEGER NOT NULL,
                medico_id INTEGER NOT NULL,
                paciente_id INTEGER NOT NULL,
                data DATETIME DEFAULT CURRENT_TIMESTAMP,
                data_dia DATE DEFAULT (DATE(CURRENT_TIMESTAMP)),
                valor_total REAL NOT NULL,
                valor_medico REAL NOT NULL,
                valor_admin REAL NOT NULL,
                status TEXT DEFAULT 'pendente' CHECK(status IN ('pendente', 'pago')),
                FOREIGN KEY (evolucao_id) REFERENCES evolucoes(id),
                FOREIGN KEY (sala_id) REFERENCES salas(id),
                FOREIGN KEY (medico_id) REFERENCES medicos(id),
                FOREIGN KEY (paciente_id) REFERENCES pacientes(id)
            )''')
            
            # Copiar dados existentes (sem UNIQUE ainda, pode haver duplicatas)
            cur.execute('''INSERT INTO consultas_realizadas_new 
                (id, evolucao_id, sala_id, medico_id, paciente_id, data, data_dia, 
                 valor_total, valor_medico, valor_admin, status)
                SELECT id, evolucao_id, sala_id, medico_id, paciente_id, data, DATE(data),
                       valor_total, valor_medico, valor_admin, status
                FROM consultas_realizadas''')
            
            # Remover tabela antiga e renomear
            cur.execute('DROP TABLE consultas_realizadas')
            cur.execute('ALTER TABLE consultas_realizadas_new RENAME TO consultas_realizadas')
            
            # Criar índice único (ignorando conflitos de dados antigos)
            try:
                cur.execute('CREATE UNIQUE INDEX idx_sala_data_dia ON consultas_realizadas(sala_id, data_dia)')
            except:
                # Se falhar devido a duplicatas, criar índice não-único
                cur.execute('CREATE INDEX idx_sala_data_dia ON consultas_realizadas(sala_id, data_dia)')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS sistema_config (
        id INTEGER PRIMARY KEY,
        platform_name TEXT DEFAULT 'MedConnect',
        support_email TEXT DEFAULT 'suporte@medconnect.com',
        support_phone TEXT DEFAULT '(11) 3000-0000',
        default_consultation_value REAL DEFAULT 150.00,
        default_doctor_percentage REAL DEFAULT 70.00,
        default_admin_percentage REAL DEFAULT 30.00,
        require_strong_password BOOLEAN DEFAULT 0,
        two_factor_auth BOOLEAN DEFAULT 0,
        activity_log BOOLEAN DEFAULT 0,
        stun_server TEXT DEFAULT 'stun:stun.l.google.com:19302',
        allow_recording BOOLEAN DEFAULT 0,
        default_video_quality TEXT DEFAULT 'high',
        consultation_reminder BOOLEAN DEFAULT 0,
        push_notifications BOOLEAN DEFAULT 0
    )''')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS fechamentos_mensais (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        medico_id INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        ano INTEGER NOT NULL,
        total_consultas INTEGER DEFAULT 0,
        valor_total REAL DEFAULT 0,
        valor_medico REAL DEFAULT 0,
        valor_admin REAL DEFAULT 0,
        data_fechamento DATETIME DEFAULT CURRENT_TIMESTAMP,
        pagamento_confirmado_admin BOOLEAN DEFAULT 0,
        data_confirmacao_admin DATETIME,
        observacoes_admin TEXT,
        pagamento_recebido_medico BOOLEAN DEFAULT 0,
        data_confirmacao_medico DATETIME,
        observacoes_medico TEXT,
        FOREIGN KEY (medico_id) REFERENCES medicos(id),
        UNIQUE(medico_id, mes, ano)
    )''')
    
    try:
        cur.execute("ALTER TABLE users ADD COLUMN profile_photo TEXT DEFAULT NULL")
    except sqlite3.OperationalError:
        pass
    
    try:
        cur.execute("ALTER TABLE medicos ADD COLUMN chave_pix TEXT DEFAULT NULL")
    except sqlite3.OperationalError:
        pass
    
    try:
        cur.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='evolucoes'")
        table_sql = cur.fetchone()
        if table_sql and 'sala_id INTEGER NOT NULL' in table_sql[0]:
            cur.execute("PRAGMA foreign_keys=off")
            
            cur.execute('''CREATE TABLE evolucoes_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sala_id INTEGER,
                medico_id INTEGER NOT NULL,
                paciente_id INTEGER NOT NULL,
                data DATETIME DEFAULT CURRENT_TIMESTAMP,
                anotacoes TEXT,
                diagnostico TEXT,
                prescricao TEXT,
                FOREIGN KEY (sala_id) REFERENCES salas(id),
                FOREIGN KEY (medico_id) REFERENCES medicos(id),
                FOREIGN KEY (paciente_id) REFERENCES pacientes(id)
            )''')
            
            cur.execute('''INSERT INTO evolucoes_new (id, sala_id, medico_id, paciente_id, data, anotacoes, diagnostico, prescricao)
                SELECT id, sala_id, medico_id, paciente_id, data, anotacoes, diagnostico, prescricao 
                FROM evolucoes''')
            
            cur.execute("DROP TABLE evolucoes")
            cur.execute("ALTER TABLE evolucoes_new RENAME TO evolucoes")
            
            cur.execute("PRAGMA foreign_keys=on")
            conn.commit()
    except (sqlite3.OperationalError, sqlite3.IntegrityError):
        pass
    
    cur.execute("SELECT COUNT(*) as count FROM users WHERE tipo='admin'")
    admin_count = cur.fetchone()[0]
    if admin_count == 0:
        admin_password = os.environ.get('ADMIN_PASSWORD')
        if not admin_password:
            alphabet = string.ascii_letters + string.digits + '!@#$%&*'
            admin_password = ''.join(secrets.choice(alphabet) for _ in range(16))
            print('=' * 80)
            print('ATENÇÃO: Senha do admin gerada automaticamente!')
            print(f'Email: admin@consultas.com')
            print(f'Senha: {admin_password}')
            print('GUARDE ESTA SENHA EM LOCAL SEGURO!')
            print('Para definir uma senha customizada, configure a variável ADMIN_PASSWORD')
            print('=' * 80)
        senha_hash = generate_password_hash(admin_password)
        cur.execute("INSERT INTO users (email, senha_hash, nome, tipo) VALUES (?, ?, ?, ?)",
                   ('admin@consultas.com', senha_hash, 'Administrador', 'admin'))
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_users_tipo ON users(tipo)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_medicos_user_id ON medicos(user_id)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_pacientes_user_id ON pacientes(user_id)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_pacientes_medico_id ON pacientes(medico_id)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_salas_medico_id ON salas(medico_id)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_salas_paciente_id ON salas(paciente_id)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_salas_ativa ON salas(ativa)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_salas_codigo ON salas(codigo)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_evolucoes_medico_id ON evolucoes(medico_id)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_evolucoes_paciente_id ON evolucoes(paciente_id)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_evolucoes_sala_id ON evolucoes(sala_id)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_consultas_medico_id ON consultas_realizadas(medico_id)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_consultas_paciente_id ON consultas_realizadas(paciente_id)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_consultas_data ON consultas_realizadas(data)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_consultas_data_dia ON consultas_realizadas(data_dia)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_fechamentos_medico_mes_ano ON fechamentos_mensais(medico_id, mes, ano)')
    except:
        pass
    
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_faturamento_config_medico_id ON faturamento_config(medico_id)')
    except:
        pass
    
    conn.commit()
    conn.close()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Você precisa fazer login primeiro.')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Você precisa fazer login primeiro.')
                return redirect(url_for('login'))
            if session.get('tipo') not in roles:
                flash('Você não tem permissão para acessar esta página.')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def gerar_codigo(n=6):
    chars = string.ascii_uppercase + string.digits
    return ''.join(random.choice(chars) for _ in range(n))

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        senha = request.form.get('senha')
        
        if not email or not senha:
            flash('Email e senha são obrigatórios.')
            return render_template('login.html')
        
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
        conn.close()
        
        if user and check_password_hash(user['senha_hash'], senha):
            session['user_id'] = user['id']
            session['nome'] = user['nome']
            session['tipo'] = user['tipo']
            session['email'] = user['email']
            flash(f'Bem-vindo(a), {user["nome"]}!')
            return redirect(url_for('dashboard'))
        else:
            flash('Email ou senha incorretos.')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Você saiu do sistema.')
    return redirect(url_for('login'))

@app.route('/perfil')
@login_required
def perfil():
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    
    medico_info = None
    if user['tipo'] == 'medico':
        medico_info = conn.execute('SELECT * FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    conn.close()
    return render_template('perfil.html', user=user, medico_info=medico_info)

@app.route('/upload-foto-perfil', methods=['POST'])
@login_required
def upload_foto_perfil():
    if 'foto' not in request.files:
        flash('Nenhuma foto foi selecionada.')
        return redirect(url_for('perfil'))
    
    file = request.files['foto']
    
    if file.filename == '':
        flash('Nenhuma foto foi selecionada.')
        return redirect(url_for('perfil'))
    
    if file and file.filename and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        ext = filename.rsplit('.', 1)[1].lower()
        new_filename = f"user_{session['user_id']}.{ext}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
        
        if os.path.exists(filepath):
            os.remove(filepath)
        
        file.save(filepath)
        
        conn = get_db()
        conn.execute('UPDATE users SET profile_photo = ? WHERE id = ?', 
                    (new_filename, session['user_id']))
        conn.commit()
        conn.close()
        
        flash('Foto de perfil atualizada com sucesso!')
        return redirect(url_for('perfil'))
    else:
        flash('Tipo de arquivo não permitido. Use: PNG, JPG, JPEG, GIF ou WEBP.')
        return redirect(url_for('perfil'))

@app.route('/atualizar-chave-pix', methods=['POST'])
@role_required('medico')
def atualizar_chave_pix():
    chave_pix = request.form.get('chave_pix', '').strip()
    
    conn = get_db()
    medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    if not medico:
        flash('Médico não encontrado.', 'error')
        conn.close()
        return redirect(url_for('perfil'))
    
    try:
        conn.execute('UPDATE medicos SET chave_pix = ? WHERE id = ?', (chave_pix, medico['id']))
        conn.commit()
        flash('Chave PIX atualizada com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao atualizar chave PIX: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('perfil'))

@app.route('/trocar-senha', methods=['POST'])
@login_required
def trocar_senha():
    senha_atual = request.form.get('senha_atual')
    nova_senha = request.form.get('nova_senha')
    confirmar_senha = request.form.get('confirmar_senha')
    
    if not all([senha_atual, nova_senha, confirmar_senha]):
        flash('Todos os campos são obrigatórios.', 'error')
        return redirect(url_for('perfil'))
    
    if nova_senha != confirmar_senha:
        flash('A nova senha e a confirmação não coincidem.', 'error')
        return redirect(url_for('perfil'))
    
    if len(nova_senha) < 4:
        flash('A nova senha deve ter pelo menos 4 caracteres.', 'error')
        return redirect(url_for('perfil'))
    
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    
    if not user or not check_password_hash(user['senha_hash'], senha_atual):
        conn.close()
        flash('Senha atual incorreta.', 'error')
        return redirect(url_for('perfil'))
    
    try:
        nova_senha_hash = generate_password_hash(str(nova_senha))
        conn.execute('UPDATE users SET senha_hash = ? WHERE id = ?', (nova_senha_hash, session['user_id']))
        conn.commit()
        flash('Senha alterada com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao alterar senha: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('perfil'))

@app.route('/admin/resetar-senha/<int:user_id>', methods=['POST'])
@role_required('admin')
def admin_resetar_senha(user_id):
    conn = get_db()
    user = conn.execute('SELECT nome, email FROM users WHERE id = ?', (user_id,)).fetchone()
    
    if not user:
        flash('Usuário não encontrado.', 'error')
        conn.close()
        return redirect(url_for('admin_dashboard'))
    
    try:
        senha_padrao = '123456'
        senha_hash = generate_password_hash(senha_padrao)
        conn.execute('UPDATE users SET senha_hash = ? WHERE id = ?', (senha_hash, user_id))
        conn.commit()
        flash(f'Senha do usuário {user["nome"]} resetada para 123456 com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao resetar senha: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin_dashboard'))

@app.route('/dashboard')
@login_required
def dashboard():
    tipo = session.get('tipo')
    if tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    elif tipo == 'medico':
        return redirect(url_for('medico_dashboard'))
    elif tipo == 'paciente':
        return redirect(url_for('paciente_dashboard'))
    
    return redirect(url_for('login'))

@app.route('/principal')
@role_required('admin', 'medico')
def dashboard_principal():
    conn = get_db()
    tipo = session.get('tipo')
    
    if tipo == 'admin':
        # Buscar dados para admin
        medicos = conn.execute('''
            SELECT m.id as medico_id, u.nome, u.email, m.crm, m.especialidade, u.profile_photo 
            FROM users u 
            LEFT JOIN medicos m ON u.id = m.user_id 
            WHERE u.tipo = "medico"
        ''').fetchall()
        
        pacientes = conn.execute('''
            SELECT u.id, u.nome, u.email, p.data_nascimento, p.telefone,
                   um.nome as medico_nome, u.profile_photo
            FROM users u 
            LEFT JOIN pacientes p ON u.id = p.user_id
            LEFT JOIN medicos m ON p.medico_id = m.id
            LEFT JOIN users um ON m.user_id = um.id
            WHERE u.tipo = "paciente"
        ''').fetchall()
        conn.close()
        
        return render_template('dashboard_principal.html', medicos=medicos, pacientes=pacientes, tipo=tipo)
    
    elif tipo == 'medico':
        # Buscar dados para médico
        medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
        
        if medico:
            pacientes = conn.execute('''
                SELECT p.id, u.nome, u.email, p.data_nascimento, p.telefone,
                       s.codigo as sala_codigo, s.id as sala_id, u.profile_photo
                FROM pacientes p
                JOIN users u ON p.user_id = u.id
                LEFT JOIN salas s ON s.paciente_id = p.id AND s.medico_id = ? AND s.ativa = 1
                WHERE p.medico_id = ?
            ''', (medico['id'], medico['id'])).fetchall()
        else:
            pacientes = []
            flash('Perfil de médico não encontrado. Entre em contato com o administrador.')
        
        medicos = []
        conn.close()
        
        return render_template('dashboard_principal.html', medicos=medicos, pacientes=pacientes, tipo=tipo)
    
    conn.close()
    return redirect(url_for('dashboard'))

@app.route('/admin')
@role_required('admin')
def admin_dashboard():
    conn = get_db()
    medicos = conn.execute('''
        SELECT m.id as medico_id, u.id as user_id, u.nome, u.email, m.crm, m.especialidade, u.profile_photo 
        FROM users u 
        LEFT JOIN medicos m ON u.id = m.user_id 
        WHERE u.tipo = "medico"
    ''').fetchall()
    
    pacientes = conn.execute('''
        SELECT u.id as user_id, u.nome, u.email, p.data_nascimento, p.telefone,
               um.nome as medico_nome, u.profile_photo
        FROM users u 
        LEFT JOIN pacientes p ON u.id = p.user_id
        LEFT JOIN medicos m ON p.medico_id = m.id
        LEFT JOIN users um ON m.user_id = um.id
        WHERE u.tipo = "paciente"
    ''').fetchall()
    conn.close()
    
    return render_template('admin_dashboard_new.html', medicos=medicos, pacientes=pacientes)

@app.route('/admin/configuracoes', methods=['GET', 'POST'])
@role_required('admin')
def admin_configuracoes():
    conn = get_db()
    
    if request.method == 'POST':
        try:
            platform_name = request.form.get('platform_name', 'MedConnect')
            support_email = request.form.get('support_email', 'suporte@medconnect.com')
            support_phone = request.form.get('support_phone', '(11) 3000-0000')
            
            default_consultation_value = request.form.get('default_consultation_value', '')
            default_consultation_value = float(default_consultation_value) if default_consultation_value else 150.00
            
            default_doctor_percentage = request.form.get('default_doctor_percentage', '')
            default_doctor_percentage = float(default_doctor_percentage) if default_doctor_percentage else 70.00
            
            default_admin_percentage = request.form.get('default_admin_percentage', '')
            default_admin_percentage = float(default_admin_percentage) if default_admin_percentage else 30.00
            
            require_strong_password = 1 if request.form.get('require_strong_password') else 0
            two_factor_auth = 1 if request.form.get('two_factor_auth') else 0
            activity_log = 1 if request.form.get('activity_log') else 0
            stun_server = request.form.get('stun_server', 'stun:stun.l.google.com:19302')
            allow_recording = 1 if request.form.get('allow_recording') else 0
            default_video_quality = request.form.get('default_video_quality', 'high')
            consultation_reminder = 1 if request.form.get('consultation_reminder') else 0
            push_notifications = 1 if request.form.get('push_notifications') else 0
            
            cur = conn.cursor()
            existing = conn.execute('SELECT id FROM sistema_config WHERE id = 1').fetchone()
            
            if existing:
                cur.execute('''
                    UPDATE sistema_config SET
                        platform_name = ?, support_email = ?, support_phone = ?,
                        default_consultation_value = ?, default_doctor_percentage = ?, default_admin_percentage = ?,
                        require_strong_password = ?, two_factor_auth = ?, activity_log = ?,
                        stun_server = ?, allow_recording = ?, default_video_quality = ?,
                        consultation_reminder = ?, push_notifications = ?
                    WHERE id = 1
                ''', (platform_name, support_email, support_phone,
                      default_consultation_value, default_doctor_percentage, default_admin_percentage,
                      require_strong_password, two_factor_auth, activity_log,
                      stun_server, allow_recording, default_video_quality,
                      consultation_reminder, push_notifications))
            else:
                cur.execute('''
                    INSERT INTO sistema_config (
                        id, platform_name, support_email, support_phone,
                        default_consultation_value, default_doctor_percentage, default_admin_percentage,
                        require_strong_password, two_factor_auth, activity_log,
                        stun_server, allow_recording, default_video_quality,
                        consultation_reminder, push_notifications
                    ) VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (platform_name, support_email, support_phone,
                      default_consultation_value, default_doctor_percentage, default_admin_percentage,
                      require_strong_password, two_factor_auth, activity_log,
                      stun_server, allow_recording, default_video_quality,
                      consultation_reminder, push_notifications))
        except ValueError as e:
            conn.close()
            flash(f'Erro: Valores numéricos inválidos. Por favor, verifique os campos de faturamento.')
            return redirect(url_for('admin_configuracoes'))
        
        conn.commit()
        conn.close()
        
        global _system_config_cache, _config_cache_time
        _system_config_cache = None
        _config_cache_time = None
        
        flash('Configurações salvas com sucesso!')
        return redirect(url_for('admin_configuracoes'))
    
    config_row = conn.execute('SELECT * FROM sistema_config WHERE id = 1').fetchone()
    conn.close()
    
    config = dict(config_row) if config_row else {}
    
    return render_template('admin_configuracoes.html', config=config)

@app.route('/admin/cadastrar-medico', methods=['POST'])
@role_required('admin')
def cadastrar_medico():
    nome = request.form.get('nome')
    email = request.form.get('email')
    senha = request.form.get('senha')
    crm = request.form.get('crm')
    especialidade = request.form.get('especialidade')
    
    if not all([nome, email, senha]):
        flash('Nome, email e senha são obrigatórios.')
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db()
    try:
        senha_hash = generate_password_hash(str(senha))
        cur = conn.cursor()
        cur.execute('INSERT INTO users (email, senha_hash, nome, tipo) VALUES (?, ?, ?, ?)',
                   (email, senha_hash, nome, 'medico'))
        user_id = cur.lastrowid
        cur.execute('INSERT INTO medicos (user_id, crm, especialidade) VALUES (?, ?, ?)',
                   (user_id, crm, especialidade))
        conn.commit()
        flash(f'Médico {nome} cadastrado com sucesso!')
    except sqlite3.IntegrityError:
        flash('Este email já está cadastrado.')
    finally:
        conn.close()
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/cadastrar-paciente', methods=['POST'])
@role_required('admin')
def admin_cadastrar_paciente():
    nome = request.form.get('nome')
    email = request.form.get('email')
    senha = request.form.get('senha')
    data_nascimento = request.form.get('data_nascimento')
    telefone = request.form.get('telefone')
    medico_id = request.form.get('medico_id')
    
    if not all([nome, email, senha]):
        flash('Nome, email e senha são obrigatórios.')
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db()
    try:
        senha_hash = generate_password_hash(str(senha))
        cur = conn.cursor()
        cur.execute('INSERT INTO users (email, senha_hash, nome, tipo) VALUES (?, ?, ?, ?)',
                   (email, senha_hash, nome, 'paciente'))
        user_id = cur.lastrowid
        cur.execute('INSERT INTO pacientes (user_id, data_nascimento, telefone, medico_id) VALUES (?, ?, ?, ?)',
                   (user_id, data_nascimento, telefone, medico_id if medico_id else None))
        conn.commit()
        flash(f'Paciente {nome} cadastrado com sucesso!')
    except sqlite3.IntegrityError:
        flash('Este email já está cadastrado.')
    finally:
        conn.close()
    
    return redirect(url_for('admin_dashboard'))

@app.route('/medico')
@role_required('medico')
def medico_dashboard():
    conn = get_db()
    medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    pacientes = conn.execute('''
        SELECT p.id, u.nome, u.email, p.data_nascimento, p.telefone,
               s.codigo as sala_codigo, s.id as sala_id, u.profile_photo
        FROM pacientes p
        JOIN users u ON p.user_id = u.id
        LEFT JOIN salas s ON s.paciente_id = p.id AND s.medico_id = ? AND s.ativa = 1
        WHERE p.medico_id = ?
    ''', (medico['id'], medico['id'])).fetchall()
    conn.close()
    
    return render_template('medico_dashboard.html', pacientes=pacientes)

@app.route('/medico/cadastrar-paciente', methods=['POST'])
@role_required('medico')
def medico_cadastrar_paciente():
    nome = request.form.get('nome')
    email = request.form.get('email')
    senha = request.form.get('senha')
    data_nascimento = request.form.get('data_nascimento')
    telefone = request.form.get('telefone')
    
    if not all([nome, email, senha]):
        flash('Nome, email e senha são obrigatórios.')
        return redirect(url_for('medico_dashboard'))
    
    conn = get_db()
    medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    try:
        senha_hash = generate_password_hash(str(senha))
        cur = conn.cursor()
        cur.execute('INSERT INTO users (email, senha_hash, nome, tipo) VALUES (?, ?, ?, ?)',
                   (email, senha_hash, nome, 'paciente'))
        user_id = cur.lastrowid
        cur.execute('INSERT INTO pacientes (user_id, data_nascimento, telefone, medico_id) VALUES (?, ?, ?, ?)',
                   (user_id, data_nascimento, telefone, medico['id']))
        conn.commit()
        flash(f'Paciente {nome} cadastrado com sucesso!')
    except sqlite3.IntegrityError:
        flash('Este email já está cadastrado.')
    finally:
        conn.close()
    
    return redirect(url_for('medico_dashboard'))

@app.route('/medico/criar-sala/<int:paciente_id>')
@role_required('medico')
def criar_sala_paciente(paciente_id):
    conn = get_db()
    medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    if not medico:
        conn.close()
        flash('Erro: perfil de médico não encontrado.')
        return redirect(url_for('medico_dashboard'))
    
    paciente = conn.execute('SELECT u.nome FROM pacientes p JOIN users u ON p.user_id = u.id WHERE p.id = ?', 
                           (paciente_id,)).fetchone()
    
    if not paciente:
        conn.close()
        flash('Erro: paciente não encontrado.')
        return redirect(url_for('medico_dashboard'))
    
    sala_existente = conn.execute(
        'SELECT codigo, ativa FROM salas WHERE medico_id = ? AND paciente_id = ?',
        (medico['id'], paciente_id)
    ).fetchone()
    
    if sala_existente:
        if sala_existente['ativa'] == 0:
            cur = conn.cursor()
            cur.execute('UPDATE salas SET ativa = 1 WHERE codigo = ?', (sala_existente['codigo'],))
            conn.commit()
            flash('Sala reativada com sucesso!')
        conn.close()
        return redirect(url_for('sala', codigo=sala_existente['codigo']))
    
    codigo = gerar_codigo(6)
    
    cur = conn.cursor()
    cur.execute(
        'INSERT INTO salas (codigo, medico_id, paciente_id, titulo) VALUES (?, ?, ?, ?)',
        (codigo, medico['id'], paciente_id, f"Consulta - {paciente['nome']}")
    )
    conn.commit()
    conn.close()
    
    return redirect(url_for('sala', codigo=codigo))

@app.route('/paciente')
@role_required('paciente')
def paciente_dashboard():
    conn = get_db()
    paciente = conn.execute('SELECT id, medico_id FROM pacientes WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    salas = conn.execute('''
        SELECT s.codigo, s.titulo, s.criado_em, u.nome as medico_nome, u.profile_photo
        FROM salas s
        JOIN medicos m ON s.medico_id = m.id
        JOIN users u ON m.user_id = u.id
        WHERE s.paciente_id = ? AND s.ativa = 1
    ''', (paciente['id'],)).fetchall()
    
    evolucoes = conn.execute('''
        SELECT e.data, e.anotacoes, e.diagnostico, e.prescricao, u.nome as medico_nome
        FROM evolucoes e
        JOIN medicos m ON e.medico_id = m.id
        JOIN users u ON m.user_id = u.id
        WHERE e.paciente_id = ?
        ORDER BY e.data DESC
    ''', (paciente['id'],)).fetchall()
    
    conn.close()
    
    return render_template('paciente_dashboard.html', salas=salas, evolucoes=evolucoes)

@app.route('/paciente/evolucoes')
@role_required('paciente')
def paciente_evolucoes():
    conn = get_db()
    paciente = conn.execute('SELECT id, medico_id FROM pacientes WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    if not paciente:
        flash('Paciente não encontrado.')
        conn.close()
        return redirect(url_for('dashboard'))
    
    evolucoes = conn.execute('''
        SELECT e.id, e.data, e.anotacoes, e.diagnostico, e.prescricao, 
               u.nome as medico_nome, u.profile_photo, m.especialidade,
               s.titulo as sala_titulo
        FROM evolucoes e
        LEFT JOIN medicos m ON e.medico_id = m.id
        LEFT JOIN users u ON m.user_id = u.id
        LEFT JOIN salas s ON e.sala_id = s.id
        WHERE e.paciente_id = ?
        ORDER BY e.data DESC
    ''', (paciente['id'],)).fetchall()
    
    conn.close()
    
    return render_template('paciente_evolucoes.html', evolucoes=evolucoes)

@app.route('/sala/<codigo>')
@login_required
def sala(codigo):
    conn = get_db()
    sala = conn.execute('SELECT * FROM salas WHERE codigo = ?', (codigo,)).fetchone()
    
    if not sala:
        flash('Sala não encontrada.')
        conn.close()
        return redirect(url_for('dashboard'))
    
    medico = conn.execute('''
        SELECT u.nome FROM medicos m 
        JOIN users u ON m.user_id = u.id 
        WHERE m.id = ?
    ''', (sala['medico_id'],)).fetchone()
    
    paciente = conn.execute('''
        SELECT u.nome FROM pacientes p 
        JOIN users u ON p.user_id = u.id 
        WHERE p.id = ?
    ''', (sala['paciente_id'],)).fetchone()
    
    conn.close()
    
    return render_template('sala.html', 
                         codigo=codigo, 
                         sala_id=sala['id'],
                         medico_nome=medico['nome'],
                         paciente_nome=paciente['nome'],
                         is_medico=session.get('tipo') == 'medico')

@app.route('/sala/<codigo>/evolucao', methods=['POST'])
@role_required('medico')
def criar_evolucao(codigo):
    conn = get_db()
    conn.execute('BEGIN IMMEDIATE')
    
    try:
        sala = conn.execute('SELECT * FROM salas WHERE codigo = ?', (codigo,)).fetchone()
        medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
        
        if sala and medico and sala['medico_id'] == medico['id']:
            anotacoes = request.form.get('anotacoes')
            diagnostico = request.form.get('diagnostico')
            prescricao = request.form.get('prescricao')
            
            cur = conn.cursor()
            cur.execute('''
                INSERT INTO evolucoes (sala_id, medico_id, paciente_id, anotacoes, diagnostico, prescricao)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (sala['id'], medico['id'], sala['paciente_id'], anotacoes, diagnostico, prescricao))
            evolucao_id = cur.lastrowid
            
            # Data de hoje para garantir consistência
            hoje = datetime.now().date().isoformat()
            
            # Verificar se já existe consulta registrada hoje para esta sala
            consulta_existente = conn.execute('''
                SELECT id FROM consultas_realizadas
                WHERE sala_id = ? AND data_dia = ?
            ''', (sala['id'], hoje)).fetchone()
            
            if consulta_existente:
                # Se já existe consulta, apenas vincular a evolução
                cur.execute('''
                    UPDATE consultas_realizadas 
                    SET evolucao_id = ?
                    WHERE id = ?
                ''', (evolucao_id, consulta_existente['id']))
                flash('Evolução registrada e vinculada à consulta!')
            else:
                # Se não existe, criar nova consulta
                faturamento = conn.execute('SELECT * FROM faturamento_config WHERE medico_id = ?', (medico['id'],)).fetchone()
                
                if not faturamento:
                    cur.execute('INSERT INTO faturamento_config (medico_id) VALUES (?)', (medico['id'],))
                    valor_consulta = 150.00
                    porcentagem_medico = 70.00
                    porcentagem_admin = 30.00
                else:
                    valor_consulta = faturamento['valor_consulta']
                    porcentagem_medico = faturamento['porcentagem_medico']
                    porcentagem_admin = faturamento['porcentagem_admin']
                
                valor_medico = round((valor_consulta * porcentagem_medico) / 100, 2)
                valor_admin = round((valor_consulta * porcentagem_admin) / 100, 2)
                
                try:
                    cur.execute('''
                        INSERT INTO consultas_realizadas 
                        (evolucao_id, sala_id, medico_id, paciente_id, data_dia, valor_total, valor_medico, valor_admin)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (evolucao_id, sala['id'], medico['id'], sala['paciente_id'], hoje, valor_consulta, valor_medico, valor_admin))
                    
                    flash('Evolução registrada e consulta contabilizada!')
                except sqlite3.IntegrityError:
                    # Race condition: consulta foi criada entre a verificação e o INSERT
                    # Vincular a evolução à consulta existente
                    consulta_existente = conn.execute('''
                        SELECT id FROM consultas_realizadas
                        WHERE sala_id = ? AND data_dia = ?
                    ''', (sala['id'], hoje)).fetchone()
                    
                    cur.execute('''
                        UPDATE consultas_realizadas 
                        SET evolucao_id = ?
                        WHERE id = ?
                    ''', (evolucao_id, consulta_existente['id']))
                    
                    flash('Evolução registrada e vinculada à consulta!')
            
            conn.commit()
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao registrar evolução: {str(e)}')
    finally:
        conn.close()
        
    return redirect(url_for('sala', codigo=codigo))

@app.route('/sala/<codigo>/finalizar', methods=['POST'])
@role_required('medico')
def finalizar_consulta(codigo):
    conn = get_db()
    sala = conn.execute('SELECT * FROM salas WHERE codigo = ?', (codigo,)).fetchone()
    medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    if sala and medico and sala['medico_id'] == medico['id']:
        conn.execute('UPDATE salas SET ativa = 0 WHERE codigo = ?', (codigo,))
        conn.commit()
        flash('Consulta finalizada com sucesso! A sala foi desativada.')
    else:
        flash('Você não tem permissão para finalizar esta consulta.')
    
    conn.close()
    return redirect(url_for('medico_dashboard'))

@app.route('/paciente/<int:paciente_id>/detalhes')
@role_required('medico')
def paciente_detalhes(paciente_id):
    conn = get_db()
    medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    if not medico:
        flash('Erro ao identificar médico.')
        conn.close()
        return redirect(url_for('medico_dashboard'))
    
    paciente = conn.execute('''
        SELECT p.id, u.nome, u.email, p.data_nascimento, p.telefone, u.profile_photo
        FROM pacientes p
        JOIN users u ON p.user_id = u.id
        WHERE p.id = ? AND p.medico_id = ?
    ''', (paciente_id, medico['id'])).fetchone()
    
    if not paciente:
        flash('Paciente não encontrado ou você não tem permissão para visualizar.')
        conn.close()
        return redirect(url_for('medico_dashboard'))
    
    evolucoes = conn.execute('''
        SELECT e.id, e.data, e.anotacoes, e.diagnostico, e.prescricao, 
               u.nome as medico_nome, s.titulo as sala_titulo
        FROM evolucoes e
        JOIN medicos m ON e.medico_id = m.id
        JOIN users u ON m.user_id = u.id
        LEFT JOIN salas s ON e.sala_id = s.id
        WHERE e.paciente_id = ?
        ORDER BY e.data DESC
    ''', (paciente_id,)).fetchall()
    
    conn.close()
    
    return render_template('paciente_detalhes.html', paciente=paciente, evolucoes=evolucoes)

@app.route('/paciente/<int:paciente_id>/evolucao', methods=['POST'])
@role_required('medico')
def criar_evolucao_paciente(paciente_id):
    conn = get_db()
    medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    if not medico:
        flash('Erro ao identificar médico.')
        conn.close()
        return redirect(url_for('medico_dashboard'))
    
    paciente = conn.execute('SELECT id FROM pacientes WHERE id = ? AND medico_id = ?', 
                           (paciente_id, medico['id'])).fetchone()
    
    if not paciente:
        flash('Paciente não encontrado ou você não tem permissão.')
        conn.close()
        return redirect(url_for('medico_dashboard'))
    
    anotacoes = request.form.get('anotacoes')
    diagnostico = request.form.get('diagnostico')
    prescricao = request.form.get('prescricao')
    
    if not anotacoes:
        flash('As anotações são obrigatórias.')
        conn.close()
        return redirect(url_for('paciente_detalhes', paciente_id=paciente_id))
    
    cur = conn.cursor()
    cur.execute('''
        INSERT INTO evolucoes (sala_id, medico_id, paciente_id, anotacoes, diagnostico, prescricao)
        VALUES (NULL, ?, ?, ?, ?, ?)
    ''', (medico['id'], paciente_id, anotacoes, diagnostico, prescricao))
    conn.commit()
    conn.close()
    
    flash('Evolução registrada com sucesso!')
    return redirect(url_for('paciente_detalhes', paciente_id=paciente_id))

@app.route('/admin/faturamento')
@role_required('admin')
def admin_faturamento():
    conn = get_db()
    
    medicos_faturamento = conn.execute('''
        SELECT m.id as medico_id, u.nome, f.valor_consulta, f.porcentagem_medico, f.porcentagem_admin
        FROM medicos m
        JOIN users u ON m.user_id = u.id
        LEFT JOIN faturamento_config f ON m.id = f.medico_id
    ''').fetchall()
    
    consultas = conn.execute('''
        SELECT c.*, u.nome as medico_nome, up.nome as paciente_nome
        FROM consultas_realizadas c
        JOIN medicos m ON c.medico_id = m.id
        JOIN users u ON m.user_id = u.id
        JOIN pacientes p ON c.paciente_id = p.id
        JOIN users up ON p.user_id = up.id
        ORDER BY c.data DESC
    ''').fetchall()
    
    totais = conn.execute('''
        SELECT 
            SUM(valor_total) as total_geral,
            SUM(valor_medico) as total_medicos,
            SUM(valor_admin) as total_admin,
            COUNT(*) as total_consultas
        FROM consultas_realizadas
    ''').fetchone()
    
    conn.close()
    
    return render_template('admin_faturamento.html', 
                         medicos=medicos_faturamento,
                         consultas=consultas,
                         totais=totais)

@app.route('/admin/faturamento/configurar/<int:medico_id>', methods=['POST'])
@role_required('admin')
def configurar_faturamento(medico_id):
    valor_consulta = request.form.get('valor_consulta', type=float)
    porcentagem_medico = request.form.get('porcentagem_medico', type=float)
    porcentagem_admin = request.form.get('porcentagem_admin', type=float)
    
    if valor_consulta is None or valor_consulta <= 0:
        flash('O valor da consulta deve ser maior que zero')
        return redirect(url_for('admin_faturamento'))
    
    if porcentagem_medico is None or porcentagem_medico < 0 or porcentagem_medico > 100:
        flash('A porcentagem do médico deve estar entre 0% e 100%')
        return redirect(url_for('admin_faturamento'))
    
    if porcentagem_admin is None or porcentagem_admin < 0 or porcentagem_admin > 100:
        flash('A porcentagem do admin deve estar entre 0% e 100%')
        return redirect(url_for('admin_faturamento'))
    
    soma_porcentagens = round(porcentagem_medico + porcentagem_admin, 2)
    if abs(soma_porcentagens - 100) > 0.01:
        flash('A soma das porcentagens deve ser 100%')
        return redirect(url_for('admin_faturamento'))
    
    valor_consulta = round(valor_consulta, 2)
    porcentagem_medico = round(porcentagem_medico, 2)
    porcentagem_admin = round(porcentagem_admin, 2)
    
    conn = get_db()
    cur = conn.cursor()
    
    existing = conn.execute('SELECT id FROM faturamento_config WHERE medico_id = ?', (medico_id,)).fetchone()
    
    if existing:
        cur.execute('''
            UPDATE faturamento_config 
            SET valor_consulta = ?, porcentagem_medico = ?, porcentagem_admin = ?
            WHERE medico_id = ?
        ''', (valor_consulta, porcentagem_medico, porcentagem_admin, medico_id))
    else:
        cur.execute('''
            INSERT INTO faturamento_config (medico_id, valor_consulta, porcentagem_medico, porcentagem_admin)
            VALUES (?, ?, ?, ?)
        ''', (medico_id, valor_consulta, porcentagem_medico, porcentagem_admin))
    
    conn.commit()
    conn.close()
    flash('Configuração de faturamento atualizada!')
    return redirect(url_for('admin_faturamento'))

@app.route('/medico/faturamento')
@role_required('medico')
def medico_faturamento():
    conn = get_db()
    medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    
    where_clause = 'WHERE c.medico_id = ?'
    params = [medico['id']]
    
    if data_inicio and data_fim:
        where_clause += ' AND c.data BETWEEN ? AND ?'
        params.extend([data_inicio + ' 00:00:00', data_fim + ' 23:59:59'])
    elif data_inicio:
        where_clause += ' AND c.data >= ?'
        params.append(data_inicio + ' 00:00:00')
    elif data_fim:
        where_clause += ' AND c.data <= ?'
        params.append(data_fim + ' 23:59:59')
    
    config = conn.execute('SELECT * FROM faturamento_config WHERE medico_id = ?', (medico['id'],)).fetchone()
    
    consultas = conn.execute(f'''
        SELECT c.*, u.nome as paciente_nome, c.data
        FROM consultas_realizadas c
        JOIN pacientes p ON c.paciente_id = p.id
        JOIN users u ON p.user_id = u.id
        {where_clause}
        ORDER BY c.data DESC
    ''', params).fetchall()
    
    totais = conn.execute(f'''
        SELECT 
            SUM(valor_total) as total_bruto,
            SUM(valor_medico) as total_medico,
            SUM(valor_admin) as total_admin,
            COUNT(*) as total_consultas
        FROM consultas_realizadas c
        {where_clause}
    ''', params).fetchone()
    
    conn.close()
    
    return render_template('medico_faturamento.html',
                         config=config,
                         consultas=consultas,
                         totais=totais,
                         data_inicio=data_inicio,
                         data_fim=data_fim)

@app.route('/admin/relatorio-fechamento')
@role_required('admin')
def admin_relatorio_fechamento():
    conn = get_db()
    
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    
    join_condition = 'c.medico_id = m.id'
    params_medicos = []
    params_totais = []
    
    if data_inicio and data_fim:
        join_condition += ' AND c.data BETWEEN ? AND ?'
        params_medicos = [data_inicio + ' 00:00:00', data_fim + ' 23:59:59']
        params_totais = [data_inicio + ' 00:00:00', data_fim + ' 23:59:59']
        where_totais = 'WHERE c.data BETWEEN ? AND ?'
    elif data_inicio:
        join_condition += ' AND c.data >= ?'
        params_medicos = [data_inicio + ' 00:00:00']
        params_totais = [data_inicio + ' 00:00:00']
        where_totais = 'WHERE c.data >= ?'
    elif data_fim:
        join_condition += ' AND c.data <= ?'
        params_medicos = [data_fim + ' 23:59:59']
        params_totais = [data_fim + ' 23:59:59']
        where_totais = 'WHERE c.data <= ?'
    else:
        where_totais = ''
    
    relatorio_medicos = conn.execute(f'''
        SELECT 
            m.id as medico_id,
            u.nome as medico_nome,
            COUNT(c.id) as total_consultas,
            COALESCE(SUM(c.valor_total), 0) as total_faturado,
            COALESCE(SUM(c.valor_medico), 0) as total_medico,
            COALESCE(SUM(c.valor_admin), 0) as total_admin
        FROM medicos m
        JOIN users u ON m.user_id = u.id
        LEFT JOIN consultas_realizadas c ON {join_condition}
        GROUP BY m.id, u.nome
        ORDER BY u.nome
    ''', params_medicos).fetchall()
    
    if where_totais:
        totais_geral = conn.execute(f'''
            SELECT 
                COUNT(*) as total_consultas,
                COALESCE(SUM(valor_total), 0) as total_faturado,
                COALESCE(SUM(valor_medico), 0) as total_medicos,
                COALESCE(SUM(valor_admin), 0) as total_admin
            FROM consultas_realizadas c
            {where_totais}
        ''', params_totais).fetchone()
    else:
        totais_geral = conn.execute('''
            SELECT 
                COUNT(*) as total_consultas,
                COALESCE(SUM(valor_total), 0) as total_faturado,
                COALESCE(SUM(valor_medico), 0) as total_medicos,
                COALESCE(SUM(valor_admin), 0) as total_admin
            FROM consultas_realizadas c
        ''').fetchone()
    
    conn.close()
    
    return render_template('admin_relatorio_fechamento.html',
                         relatorio_medicos=relatorio_medicos,
                         totais_geral=totais_geral,
                         data_inicio=data_inicio,
                         data_fim=data_fim)

@app.route('/admin/zerar-banco', methods=['POST'])
@role_required('admin')
def admin_zerar_banco():
    confirmacao = request.form.get('confirmacao', '')
    if confirmacao.upper() != 'ZERAR':
        flash('Confirmação inválida. Digite ZERAR para confirmar.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('DELETE FROM fechamentos_mensais')
        cur.execute('DELETE FROM consultas_realizadas')
        cur.execute('DELETE FROM faturamento_config')
        cur.execute('DELETE FROM evolucoes')
        cur.execute('DELETE FROM salas')
        cur.execute('DELETE FROM pacientes')
        cur.execute('DELETE FROM medicos')
        cur.execute('DELETE FROM users WHERE tipo != "admin"')
        
        conn.commit()
        conn.close()
        
        flash('Banco de dados zerado com sucesso! Apenas o administrador foi mantido.', 'success')
    except Exception as e:
        flash(f'Erro ao zerar banco de dados: {str(e)}', 'error')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/fechamento/criar/<int:medico_id>', methods=['POST'])
@role_required('admin')
def admin_criar_fechamento(medico_id):
    mes = request.form.get('mes', type=int)
    ano = request.form.get('ano', type=int)
    
    if not mes or not ano:
        return jsonify({'success': False, 'message': 'Mês e ano são obrigatórios'}), 400
    
    conn = get_db()
    
    try:
        consultas = conn.execute('''
            SELECT 
                COUNT(*) as total_consultas,
                COALESCE(SUM(valor_total), 0) as valor_total,
                COALESCE(SUM(valor_medico), 0) as valor_medico,
                COALESCE(SUM(valor_admin), 0) as valor_admin
            FROM consultas_realizadas
            WHERE medico_id = ?
            AND strftime('%m', data) = ?
            AND strftime('%Y', data) = ?
        ''', (medico_id, f'{mes:02d}', str(ano))).fetchone()
        
        cur = conn.cursor()
        cur.execute('''
            INSERT OR REPLACE INTO fechamentos_mensais 
            (medico_id, mes, ano, total_consultas, valor_total, valor_medico, valor_admin, data_fechamento)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (medico_id, mes, ano, consultas['total_consultas'], 
              consultas['valor_total'], consultas['valor_medico'], 
              consultas['valor_admin'], datetime.now().isoformat()))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Fechamento criado com sucesso'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/fechamento/confirmar-pagamento/<int:fechamento_id>', methods=['POST'])
@role_required('admin')
def admin_confirmar_pagamento(fechamento_id):
    observacoes = request.form.get('observacoes', '')
    
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute('''
            UPDATE fechamentos_mensais 
            SET pagamento_confirmado_admin = 1,
                data_confirmacao_admin = ?,
                observacoes_admin = ?
            WHERE id = ?
        ''', (datetime.now().isoformat(), observacoes, fechamento_id))
        
        conn.commit()
        conn.close()
        
        flash('Pagamento confirmado com sucesso!', 'success')
        return jsonify({'success': True, 'message': 'Pagamento confirmado'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/medico/fechamento/confirmar-recebimento/<int:fechamento_id>', methods=['POST'])
@role_required('medico')
def medico_confirmar_recebimento(fechamento_id):
    observacoes = request.form.get('observacoes', '')
    
    conn = get_db()
    
    medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
    if not medico:
        conn.close()
        return jsonify({'success': False, 'message': 'Médico não encontrado'}), 404
    
    fechamento = conn.execute('SELECT * FROM fechamentos_mensais WHERE id = ? AND medico_id = ?', 
                              (fechamento_id, medico['id'])).fetchone()
    if not fechamento:
        conn.close()
        return jsonify({'success': False, 'message': 'Fechamento não encontrado'}), 404
    
    try:
        cur = conn.cursor()
        cur.execute('''
            UPDATE fechamentos_mensais 
            SET pagamento_recebido_medico = 1,
                data_confirmacao_medico = ?,
                observacoes_medico = ?
            WHERE id = ?
        ''', (datetime.now().isoformat(), observacoes, fechamento_id))
        
        conn.commit()
        conn.close()
        
        flash('Recebimento confirmado com sucesso!', 'success')
        return jsonify({'success': True, 'message': 'Recebimento confirmado'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/fechamentos')
@role_required('admin')
def admin_fechamentos():
    conn = get_db()
    
    fechamentos = conn.execute('''
        SELECT 
            f.*,
            u.nome as medico_nome
        FROM fechamentos_mensais f
        JOIN medicos m ON f.medico_id = m.id
        JOIN users u ON m.user_id = u.id
        ORDER BY f.ano DESC, f.mes DESC, u.nome
    ''').fetchall()
    
    medicos = conn.execute('''
        SELECT m.id, u.nome 
        FROM medicos m
        JOIN users u ON m.user_id = u.id
        ORDER BY u.nome
    ''').fetchall()
    
    conn.close()
    
    return render_template('admin_fechamentos.html', 
                         fechamentos=fechamentos,
                         medicos=medicos,
                         now=datetime.now())

@app.route('/medico/meus-pagamentos')
@role_required('medico')
def medico_pagamentos():
    conn = get_db()
    
    medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (session['user_id'],)).fetchone()
    if not medico:
        flash('Médico não encontrado.', 'error')
        return redirect(url_for('dashboard'))
    
    fechamentos = conn.execute('''
        SELECT * FROM fechamentos_mensais
        WHERE medico_id = ?
        ORDER BY ano DESC, mes DESC
    ''', (medico['id'],)).fetchall()
    
    conn.close()
    
    return render_template('medico_pagamentos.html', fechamentos=fechamentos)

@app.route('/admin/exportar-planilha-pagamentos')
@role_required('admin')
def exportar_planilha_pagamentos():
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)
    
    conn = get_db()
    
    if mes and ano:
        fechamentos = conn.execute('''
            SELECT 
                f.*,
                u.nome as medico_nome,
                m.chave_pix
            FROM fechamentos_mensais f
            JOIN medicos m ON f.medico_id = m.id
            JOIN users u ON m.user_id = u.id
            WHERE f.mes = ? AND f.ano = ?
            ORDER BY u.nome
        ''', (mes, ano)).fetchall()
        nome_arquivo = f'pagamentos_{mes:02d}_{ano}.xlsx'
    else:
        fechamentos = conn.execute('''
            SELECT 
                f.*,
                u.nome as medico_nome,
                m.chave_pix
            FROM fechamentos_mensais f
            JOIN medicos m ON f.medico_id = m.id
            JOIN users u ON m.user_id = u.id
            WHERE f.pagamento_confirmado_admin = 1
            ORDER BY f.ano DESC, f.mes DESC, u.nome
        ''').fetchall()
        nome_arquivo = f'pagamentos_todos.xlsx'
    
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pagamentos'
    
    header_fill = PatternFill(start_color='0891B2', end_color='0891B2', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['Médico', 'Mês/Ano', 'Total Consultas', 'Valor Bruto (R$)', 'Valor Líquido (R$)', 'Chave PIX', 'Status Pagamento']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    meses_nomes = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    
    for fechamento in fechamentos:
        mes_ano = f"{meses_nomes.get(fechamento['mes'], fechamento['mes'])}/{fechamento['ano']}"
        chave_pix = fechamento['chave_pix'] if fechamento['chave_pix'] else 'Não cadastrada'
        status = 'Pago' if fechamento['pagamento_confirmado_admin'] else 'Pendente'
        
        row = [
            fechamento['medico_nome'],
            mes_ano,
            fechamento['total_consultas'],
            f"R$ {fechamento['valor_total']:.2f}",
            f"R$ {fechamento['valor_medico']:.2f}",
            chave_pix,
            status
        ]
        ws.append(row)
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 18
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nome_arquivo
    )

@socketio.on('connect')
def handle_connect():
    print(f'Cliente conectado: {request.sid}')

@socketio.on('disconnect')
def handle_disconnect():
    print(f'Cliente desconectado: {request.sid}')
    
    for sala, users in list(room_users.items()):
        if request.sid in users:
            user_data = users[request.sid]
            del users[request.sid]
            emit('peer-left', {
                'msg': f'{user_data["nome"]} saiu.', 
                'nome': user_data['nome'], 
                'tipo': user_data['tipo']
            }, to=sala)
            if not users:
                del room_users[sala]
            break

@socketio.on('join')
def on_join(data):
    sala = data.get('sala')
    nome = data.get('nome', 'visitante')
    tipo = data.get('tipo', 'visitante')
    print(f'Join event - sala: {sala}, nome: {nome}, tipo: {tipo}')
    
    join_room(sala)
    
    if sala not in room_users:
        room_users[sala] = {}
    
    for existing_sid, existing_user in room_users[sala].items():
        emit('peer-joined', {
            'msg': f'{existing_user["nome"]} já está na sala.', 
            'nome': existing_user['nome'], 
            'tipo': existing_user['tipo']
        }, to=request.sid)
    
    room_users[sala][request.sid] = {'nome': nome, 'tipo': tipo}
    
    emit('peer-joined', {'msg': f'{nome} entrou.', 'nome': nome, 'tipo': tipo}, to=sala, include_self=False)

@socketio.on('iniciar_consulta')
def iniciar_consulta(data):
    """Registra automaticamente a consulta quando a vídeo chamada inicia"""
    print(f'Iniciar consulta event recebido: {data}')
    # Validar autenticação
    if 'user_id' not in session:
        emit('consulta_error', {'error': 'Usuário não autenticado'})
        return
    
    codigo_sala = data.get('codigo_sala')
    
    if not codigo_sala:
        emit('consulta_error', {'error': 'Código da sala não fornecido'})
        return
    
    conn = get_db()
    conn.execute('BEGIN IMMEDIATE')
    
    try:
        # Buscar informações da sala
        sala = conn.execute('''
            SELECT id, medico_id, paciente_id 
            FROM salas 
            WHERE codigo = ?
        ''', (codigo_sala,)).fetchone()
        
        if not sala:
            conn.close()
            emit('consulta_error', {'error': 'Sala não encontrada'})
            return
        
        sala_id, medico_id, paciente_id = sala['id'], sala['medico_id'], sala['paciente_id']
        
        # Validar que o usuário pertence à sala (é o médico ou o paciente)
        user_id = session['user_id']
        user_tipo = session.get('tipo')
        
        pertence_sala = False
        if user_tipo == 'medico':
            medico = conn.execute('SELECT id FROM medicos WHERE user_id = ?', (user_id,)).fetchone()
            if medico and medico['id'] == medico_id:
                pertence_sala = True
        elif user_tipo == 'paciente':
            paciente = conn.execute('SELECT id FROM pacientes WHERE user_id = ?', (user_id,)).fetchone()
            if paciente and paciente['id'] == paciente_id:
                pertence_sala = True
        
        if not pertence_sala:
            conn.close()
            emit('consulta_error', {'error': 'Você não tem permissão para esta sala'})
            return
        
        # Buscar configuração de faturamento do médico
        config = conn.execute('''
            SELECT valor_consulta, porcentagem_medico, porcentagem_admin
            FROM faturamento_config
            WHERE medico_id = ?
        ''', (medico_id,)).fetchone()
        
        # Se não houver configuração, usar valores padrão
        if config:
            valor_total = config['valor_consulta']
            porc_medico = config['porcentagem_medico']
            porc_admin = config['porcentagem_admin']
        else:
            valor_total = 150.00
            porc_medico = 70.00
            porc_admin = 30.00
        
        # Calcular valores
        valor_medico = round((valor_total * porc_medico) / 100, 2)
        valor_admin = round((valor_total * porc_admin) / 100, 2)
        
        # Data de hoje para garantir consistência
        hoje = datetime.now().date().isoformat()
        
        # Tentar registrar a consulta - a constraint UNIQUE previne duplicatas
        try:
            cursor = conn.execute('''
                INSERT INTO consultas_realizadas 
                (evolucao_id, sala_id, medico_id, paciente_id, data_dia, valor_total, valor_medico, valor_admin)
                VALUES (NULL, ?, ?, ?, ?, ?, ?, ?)
            ''', (sala_id, medico_id, paciente_id, hoje, valor_total, valor_medico, valor_admin))
            
            consulta_id = cursor.lastrowid
            conn.commit()
            
            emit('consulta_registrada', {
                'consulta_id': consulta_id,
                'valor_total': valor_total,
                'msg': 'Consulta registrada com sucesso'
            })
            
        except sqlite3.IntegrityError:
            # Consulta já existe hoje (UNIQUE constraint violated)
            conn.rollback()
            consulta_existente = conn.execute('''
                SELECT id, valor_total FROM consultas_realizadas
                WHERE sala_id = ? AND data_dia = ?
            ''', (sala_id, hoje)).fetchone()
            
            emit('consulta_registrada', {
                'consulta_id': consulta_existente['id'],
                'valor_total': consulta_existente['valor_total'],
                'msg': 'Consulta já registrada anteriormente'
            })
            
    except Exception as e:
        conn.rollback()
        emit('consulta_error', {'error': str(e)})
    finally:
        conn.close()

@socketio.on('signal')
def on_signal(data):
    sala = data.get('sala')
    signal_type = data.get('type', 'unknown')
    print(f'Signal event - sala: {sala}, type: {signal_type}')
    emit('signal', data, to=sala, include_self=False)

@socketio.on('leave')
def on_leave(data):
    sala = data.get('sala')
    nome = data.get('nome', 'visitante')
    tipo = data.get('tipo', 'visitante')
    leave_room(sala)
    
    if sala in room_users and request.sid in room_users[sala]:
        del room_users[sala][request.sid]
        if not room_users[sala]:
            del room_users[sala]
    
    emit('peer-left', {'msg': f'{nome} saiu.', 'nome': nome, 'tipo': tipo}, to=sala)

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    socketio.run(app, host='0.0.0.0', port=port, debug=False, use_reloader=False, log_output=True)
